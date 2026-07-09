# vin/admin.py
import logging
from django.contrib import admin
from django.utils.html import format_html
from .models import VinConfig, VinQueryResult
from .vin import VinConfigManager
from django.contrib.auth import get_user_model

logger = logging.getLogger(__name__)
User = get_user_model()


@admin.register(VinConfig)
class VinConfigAdmin(admin.ModelAdmin):
    """VIN查询配置管理"""
    list_display = (
        'id', 'name', 'username_preview', 'is_active', 'save_query_results', 'save_miniprogram_results', 'created_at')
    list_display_links = ('id', 'name')
    list_filter = ('is_active', 'save_query_results', 'save_miniprogram_results', 'created_at')
    search_fields = ('name', 'username')
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 20
    actions = ['activate_configs', 'deactivate_configs']

    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'username', 'password', 'is_active')
        }),
        ('数据保存控制', {
            'fields': ('save_query_results', 'save_miniprogram_results'),
            'classes': ('collapse',)
        }),
        ('时间信息', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def __init__(self, model, admin_site):
        super().__init__(model, admin_site)
        # 把 list_display 中的所有字段赋值给 list_display_links
        self.list_display_links = self.list_display

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = list(super().get_readonly_fields(request, obj))
        if not request.user.is_superuser:
            readonly_fields.extend(['username', 'password'])
        return readonly_fields

    def username_preview(self, obj):
        """用户名预览"""
        if obj.username:
            import re
            if re.match(r'^1[3-9]\d{9}$', obj.username):
                return f"{obj.username[:3]}••••{obj.username[-4:]}"
            elif '@' in obj.username:
                parts = obj.username.split('@')
                if len(parts[0]) > 2:
                    return f"{parts[0][:2]}••••@{parts[1]}"
            return f"{obj.username[:2]}••••"
        return "••••"

    username_preview.short_description = "用户名"

    def save_model(self, request, obj, form, change):
        action = "修改" if change else "创建"
        logger.info(f"管理员 {request.user} {action}了VIN配置: {obj.name}")
        super().save_model(request, obj, form, change)

    def activate_configs(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'成功激活 {updated} 个配置')

    activate_configs.short_description = "激活选中的配置"

    def deactivate_configs(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'成功禁用 {updated} 个配置')

    deactivate_configs.short_description = "禁用选中的配置"


@admin.register(VinQueryResult)
class VinQueryResultAdmin(admin.ModelAdmin):
    """VIN查询结果管理"""

    list_display = (
        'id', 'vin_code', 'user', 'brand', 'model_name', 'model_year', 'model_count',
        'processing_time', 'query_status','query_time',
    )

    list_filter = ('query_status', 'user', 'query_time')
    search_fields = ('vin_code', 'brand', 'user__username', 'model_name')
    readonly_fields = (
        'query_time', 'processing_time', 'user', 'config_used',
        'raw_response_data_preview', 'model_list_preview', 'original_epc_list_preview',
        'gonggao_list_preview', 'import_list_preview', 'original_attributes_preview',
        'statistics_display', 'has_vehicle_info_display', 'has_images_display'
    )
    list_per_page = 20
    actions = ['retry_selected_queries', 'export_selected_results', 'mark_as_success', 'mark_as_failed']
    list_display_links = ('id', 'vin_code')

    fieldsets = (
        ('基础信息', {
            'fields': (
                'vin_code', 'user','query_status',  'processing_time', 'query_time'
            )
        }),
        ('核心车辆信息', {
            'fields': (
                'brand', 'model_name', 'model_year', 'build_date'
            )
        }),
        ('统计信息', {
            'fields': ('statistics_display',),
            'classes': ('collapse',)
        }),
        ('车型列表', {
            'fields': ('model_list_preview',),
            'classes': ('collapse',)
        }),
        ('原厂EPC信息', {
            'fields': ('original_epc_list_preview',),
            'classes': ('collapse',)
        }),
        ('公告信息', {
            'fields': ('gonggao_list_preview',),
            'classes': ('collapse',)
        }),
        ('进口车型信息', {
            'fields': ('import_list_preview',),
            'classes': ('collapse',)
        }),
        ('原厂属性', {
            'fields': ('original_attributes_preview',),
            'classes': ('collapse',)
        }),
        ('原始数据', {
            'fields': ('raw_response_data',),
            'classes': ('collapse',)
        }),
    )

    def __init__(self, model, admin_site):
        super().__init__(model, admin_site)
        # 把 list_display 中的所有字段赋值给 list_display_links
        self.list_display_links = self.list_display

    def has_vehicle_info_display(self, obj):
        return obj.has_vehicle_info

    has_vehicle_info_display.boolean = True
    has_vehicle_info_display.short_description = "有车辆信息"

    def has_images_display(self, obj):
        return obj.has_images

    has_images_display.boolean = True
    has_images_display.short_description = "有图片"

    def get_fieldsets(self, request, obj=None):
        if obj is None:
            return (
                ('基础信息', {
                    'fields': ('vin_code', 'user', 'config_used')
                }),
            )
        else:
            return super().get_fieldsets(request, obj)

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = list(super().get_readonly_fields(request, obj))

        if obj is not None:
            readonly_fields.extend([
                'raw_response_data_preview', 'model_list_preview',
                'original_epc_list_preview', 'gonggao_list_preview',
                'import_list_preview', 'original_attributes_preview',
                'statistics_display', 'has_vehicle_info_display', 'has_images_display'
            ])
        else:
            readonly_fields = [field for field in readonly_fields
                               if not field.endswith('_preview') and
                               field not in ['statistics_display', 'has_vehicle_info_display', 'has_images_display']]
            readonly_fields = [field for field in readonly_fields if field != 'user']

        return readonly_fields

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)

        if obj is None:
            form.base_fields['user'].initial = request.user
            form.base_fields['user'].widget.can_add_related = False
            form.base_fields['user'].widget.can_change_related = False
            form.base_fields['user'].widget.can_delete_related = False

        return form

    def save_model(self, request, obj, form, change):
        if not change:
            obj.user = request.user
        super().save_model(request, obj, form, change)

    def raw_response_data_preview(self, obj):
        """原始响应数据预览"""
        if obj.raw_response_data:
            import json
            formatted_json = json.dumps(obj.raw_response_data, ensure_ascii=False, indent=2)
            return format_html('<pre style="max-height: 300px; overflow: auto;">{}</pre>', formatted_json)
        return "无数据"

    raw_response_data_preview.short_description = "原始响应数据预览"

    def model_list_preview(self, obj):
        """车型列表预览"""
        if obj.model_list and len(obj.model_list) > 0:
            preview = []
            for i, model in enumerate(obj.model_list[:3]):
                model_detail = model.get('Model_detail', '未知车型')
                factory = model.get('Factory', '')
                preview.append(f"{i + 1}. {model_detail} ({factory})")

            if len(obj.model_list) > 3:
                preview.append(f"... 还有 {len(obj.model_list) - 3} 个车型")

            return format_html('<br>'.join(preview))
        return "无车型数据"

    model_list_preview.short_description = "车型列表预览"

    def original_epc_list_preview(self, obj):
        """原厂EPC列表预览"""
        if obj.original_epc_list and len(obj.original_epc_list) > 0:
            preview = []
            for i, epc in enumerate(obj.original_epc_list[:2]):
                epc_id = epc.get('Epc_id', '')
                preview.append(f"{i + 1}. EPC ID: {epc_id}")

            if len(obj.original_epc_list) > 2:
                preview.append(f"... 还有 {len(obj.original_epc_list) - 2} 个EPC")

            return format_html('<br>'.join(preview))
        return "无原厂EPC数据"

    original_epc_list_preview.short_description = "原厂EPC列表预览"

    def gonggao_list_preview(self, obj):
        """公告列表预览"""
        if obj.gonggao_list and len(obj.gonggao_list) > 0:
            preview = []
            for i, gonggao in enumerate(obj.gonggao_list[:2]):
                model_name = gonggao.get('F4', '未知型号')
                preview.append(f"{i + 1}. {model_name}")

            if len(obj.gonggao_list) > 2:
                preview.append(f"... 还有 {len(obj.gonggao_list) - 2} 个公告")

            return format_html('<br>'.join(preview))
        return "无公告数据"

    gonggao_list_preview.short_description = "公告列表预览"

    def import_list_preview(self, obj):
        """进口车型列表预览"""
        if obj.import_list and len(obj.import_list) > 0:
            preview = []
            for i, import_car in enumerate(obj.import_list[:2]):
                brand = import_car.get('Make_cn', '未知品牌')
                model_name = import_car.get('Model_cn', '未知型号')
                preview.append(f"{i + 1}. {brand} - {model_name}")

            if len(obj.import_list) > 2:
                preview.append(f"... 还有 {len(obj.import_list) - 2} 个进口车型")

            return format_html('<br>'.join(preview))
        return "无进口车型数据"

    import_list_preview.short_description = "进口车型列表预览"

    def original_attributes_preview(self, obj):
        """原厂属性预览"""
        if obj.original_attributes and len(obj.original_attributes) > 0:
            preview = []
            for i, attr in enumerate(obj.original_attributes[:5]):
                col_name = attr.get('Col_name', '')
                col_value = attr.get('Col_value', '')
                if col_name and col_value:
                    preview.append(f"{col_name}: {col_value}")

            if len(obj.original_attributes) > 5:
                preview.append(f"... 还有 {len(obj.original_attributes) - 5} 个属性")

            return format_html('<br>'.join(preview))
        return "无原厂属性数据"

    original_attributes_preview.short_description = "原厂属性预览"

   def statistics_display(self, obj):
    try:
        stats = obj.get_detailed_info()
        display_lines = [
            f"车型数量: {stats.get('model_count', 0)}",
            f"原厂属性数量: {stats.get('original_attributes_count', 0)}",
            f"公告数量: {stats.get('gonggao_count', 0)}",
            f"进口车型数量: {stats.get('import_count', 0)}",
            f"原厂EPC数量: {stats.get('original_epc_count', 0)}",
            f"是否有图片: {'是' if stats.get('has_images', False) else '否'}",
            f"是否有车辆信息: {'是' if obj.has_vehicle_info else '否'}"
        ]
        # 防御性检查：如果 display_lines 为空（理论上不会发生），返回默认提示
        if not display_lines:
            return "暂无统计数据"
        return format_html('<br>'.join(display_lines))
    except Exception as e:
        logger.error(f"统计信息展示出错: {str(e)}")
        return "统计信息生成失败，请检查数据完整性"

    statistics_display.short_description = "统计信息"

    def retry_selected_queries(self, request, queryset):
        """重新执行选中的查询"""
        success_count = 0
        failed_count = 0

        for obj in queryset:
            try:
                config = obj.config_used
                if not config:
                    config = VinConfigManager.get_active_config()

                if not config:
                    logger.error(f"未找到有效的VIN配置，记录ID: {obj.id}")
                    failed_count += 1
                    continue

                logger.info(f"开始重新查询VIN: {obj.vin_code}")
                parser = VinConfigManager.query_vin_with_config(obj.vin_code, config)

                if parser.is_success:
                    all_data = self._extract_complete_data_from_parser_safe(parser)

                    # 更新所有字段
                    obj.brand = all_data.get('brand', '')
                    obj.model_year = all_data.get('model_year', '')
                    obj.build_date = all_data.get('build_date', '')

                    # 提取车型名称
                    model_list = all_data.get('model_list', [])
                    if model_list and len(model_list) > 0:
                        first_model = model_list[0]
                        obj.model_name = first_model.get('Model_detail', '') or first_model.get('Model', '')

                    # 列表数据
                    obj.model_list = all_data.get('model_list', [])
                    obj.original_epc_list = all_data.get('original_epc_list', [])
                    obj.original_attributes = all_data.get('original_attributes', [])
                    obj.gonggao_list = all_data.get('gonggao_list', [])
                    obj.import_list = all_data.get('import_list', [])
                    obj.raw_response_data = all_data.get('raw_response_data', {})

                    obj.query_status = 'success'
                    obj.error_message = None

                    obj._calculate_statistics()

                    obj.save()
                    success_count += 1
                    logger.info(f"重新查询成功并保存: VIN {obj.vin_code}")
                else:
                    obj.query_status = 'failed'
                    obj.error_message = parser.message
                    obj.raw_response_data = parser.raw_data if hasattr(parser, 'raw_data') else {}
                    obj.save()
                    failed_count += 1
                    logger.warning(f"重新查询失败: VIN {obj.vin_code}, 错误: {parser.message}")

            except Exception as e:
                logger.error(f"重新查询VIN失败: {obj.vin_code}, 错误: {str(e)}", exc_info=True)
                failed_count += 1

        self.message_user(
            request,
            f"成功重新查询 {success_count} 条记录，失败 {failed_count} 条记录"
        )

    retry_selected_queries.short_description = "重新执行选中的查询"

    def _extract_complete_data_from_parser_safe(self, parser):
        """从解析器提取完整数据 - 安全版本"""
        try:
            if not hasattr(parser, 'raw_data') or not parser.raw_data:
                return self._get_empty_data()

            data = parser.raw_data.get('data', {})
            if not data:
                return self._get_empty_data()

            # 提取车型名称
            model_name = ''
            model_list = data.get('model_list', [])
            if model_list and len(model_list) > 0:
                first_model = model_list[0]
                model_name = first_model.get('Model_detail', '') or first_model.get('Model', '')

            original_attributes = []
            original_epc_list = data.get('model_original_epc_list', [])
            if not isinstance(original_epc_list, list):
                original_epc_list = []

            for epc in original_epc_list:
                if not isinstance(epc, dict):
                    continue
                car_attrs = epc.get('CarAttributes', [])
                if not isinstance(car_attrs, list):
                    continue
                for attr in car_attrs:
                    if isinstance(attr, dict) and attr.get('Language') == 'zh':
                        original_attributes.append(attr)

            return {
                'brand': data.get('brand', ''),
                'model_year': data.get('model_year_from_vin', ''),
                'build_date': data.get('build_date', ''),
                'model_name': model_name,
                'model_list': data.get('model_list', []),
                'original_epc_list': original_epc_list,
                'original_attributes': original_attributes,
                'gonggao_list': data.get('model_gonggao_list', []),
                'import_list': data.get('model_import_list', []),
                'raw_response_data': parser.raw_data
            }
        except Exception as e:
            logger.error(f"安全提取解析器数据失败: {str(e)}")
            return self._get_empty_data()

    def mark_as_success(self, request, queryset):
        """将选中的记录标记为成功"""
        updated = queryset.update(query_status='success', error_message=None)
        self.message_user(request, f'成功标记 {updated} 条记录为成功状态')

    mark_as_success.short_description = "标记为成功"

    def mark_as_failed(self, request, queryset):
        """将选中的记录标记为失败"""
        updated = queryset.update(query_status='failed', error_message='手动标记为失败')
        self.message_user(request, f'成功标记 {updated} 条记录为失败状态')

    mark_as_failed.short_description = "标记为失败"

    def export_selected_results(self, request, queryset):
        """导出选中的查询结果"""
        try:
            from datetime import datetime
            import openpyxl
            from openpyxl.styles import Font

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"VIN查询结果_选中记录_{timestamp}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VIN查询结果"

            headers = [
                'ID', 'VIN码', '查询状态', '品牌', '车型名称', '车型年份',
                '生产日期', '车型数量', '处理时间(秒)', '查询时间', '错误信息'
            ]

            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)

            for row_num, obj in enumerate(queryset, 2):
                ws.cell(row=row_num, column=1, value=obj.id)
                ws.cell(row=row_num, column=2, value=obj.vin_code)
                ws.cell(row=row_num, column=3, value=obj.get_query_status_display())
                ws.cell(row=row_num, column=4, value=obj.brand or '')
                ws.cell(row=row_num, column=5, value=obj.model_name or '')
                ws.cell(row=row_num, column=6, value=obj.model_year or '')
                ws.cell(row=row_num, column=7, value=obj.build_date or '')
                ws.cell(row=row_num, column=8, value=obj.model_count)
                ws.cell(row=row_num, column=9, value=obj.processing_time)
                ws.cell(row=row_num, column=10, value=obj.query_time.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row_num, column=11, value=obj.error_message or '')

            for column in ws.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            from django.http import HttpResponse
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            wb.save(response)

            self.message_user(request, f"成功导出 {queryset.count()} 条VIN查询结果")
            return response

        except Exception as e:
            self.message_user(request, f"导出失败: {str(e)}", level='error')

    export_selected_results.short_description = "导出选中的查询结果"

    def _get_empty_data(self):
        """获取空数据模板"""
        return {
            'brand': '',
            'model_year': '',
            'build_date': '',
            'model_name': '',
            'model_list': [],
            'original_epc_list': [],
            'original_attributes': [],
            'gonggao_list': [],
            'import_list': [],
            'raw_response_data': {}
        }
