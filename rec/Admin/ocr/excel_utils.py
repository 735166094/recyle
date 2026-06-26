# excel_utils.py
import os
import logging
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel导出工具类"""

    # 定义所有可导出的字段及其显示名称
    FIELD_MAPPING = {
        'id': 'ID',
        'owner_name': '车主/单位',
        'identification_number': '证件号码',
        'id_card_phone_number': '手机号',
        'address': '地址',
        'vin': 'VIN码',
        'vehicle_number': '车牌号码',
        'vehicle_type': '车辆类型',
        'use_character': '使用性质',
        'brand': '品牌',
        'vehicle_model': '型号',
        'engine_no': '发动机号码',
        'approved_passengers': '核定载人数',
        'register_date': '注册日期',
        'energy_type': '燃料种类',
        'unladen_mass': '整备质量',
        'remarks': '备注',
        'match_status': '匹配状态',
        'match_score': '匹配分数',
        'matched_at': '匹配时间',
        'created_at': '创建时间'
    }

    # 默认导出的字段
    DEFAULT_FIELDS = [
        'id', 'owner_name', 'identification_number', 'id_card_phone_number', 'address', 'vin',
        'vehicle_number', 'vehicle_type', 'use_character', 'brand',
        'vehicle_model', 'engine_no', 'approved_passengers', 'register_date',
        'energy_type', 'unladen_mass', 'remarks', 'match_status'
    ]

    @staticmethod
    def export_all_scrap_car_info(filename=None, fields=None):
        """
        导出所有报废车信息到Excel - 支持字段选择

        Args:
            filename: 导出的文件名
            fields: 要导出的字段列表，如果为None则使用默认字段

        Returns:
            HttpResponse: 包含Excel文件的HTTP响应
        """
        try:
            from .models import ScrapCarInfo

            logger.info("开始导出所有报废车信息")

            # 使用优化查询
            queryset = ScrapCarInfo.objects.all().select_related(
                'vehicle_record', 'id_card_record', 'business_record'
            ).order_by('-created_at')

            logger.info(f"查询到 {queryset.count()} 条记录")

            return ExcelExporter.export_scrap_car_info(queryset, filename, fields=fields)

        except Exception as e:
            logger.error(f"导出所有报废车信息失败: {str(e)}")
            raise

    @staticmethod
    def export_scrap_car_info(records, filename=None, record_count=None, fields=None):
        """
        导出报废车信息到Excel - 支持字段选择

        Args:
            records: 记录列表或查询集
            filename: 导出的文件名
            record_count: 记录数量（可选，如果不提供则自动计算）
            fields: 要导出的字段列表，如果为None则使用默认字段
        """
        try:
            # 确定记录数量
            if record_count is not None:
                count = record_count
            elif hasattr(records, 'count'):
                count = records.count()
            else:
                count = len(records)

            logger.info(f"开始导出报废车信息，记录数: {count}")

            # 确定要导出的字段
            if fields is None:
                fields = ExcelExporter.DEFAULT_FIELDS

            # 验证字段有效性
            valid_fields = []
            for field in fields:
                if field in ExcelExporter.FIELD_MAPPING:
                    valid_fields.append(field)
                else:
                    logger.warning(f"忽略无效的导出字段: {field}")

            if not valid_fields:
                logger.error("没有有效的导出字段")
                raise ValueError("没有有效的导出字段")

            logger.info(f"导出字段: {valid_fields}")

            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "报废车信息"

            # 设置表头 - 根据选择的字段生成
            headers = [ExcelExporter.FIELD_MAPPING[field] for field in valid_fields]

            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 写入表头
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 批量写入数据
            for row_num, record in enumerate(records, 2):
                for col_num, field in enumerate(valid_fields, 1):
                    value = ExcelExporter._get_field_value(record, field)
                    ws.cell(row=row_num, column=col_num, value=value)

                # 每处理100条记录记录一次进度
                if row_num % 100 == 0:
                    logger.info(f"已处理 {row_num - 1} 条记录")

            # 自动调整列宽
            ExcelExporter._auto_adjust_columns(ws)

            # 设置文件名
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"报废车信息_{timestamp}.xlsx"

            # 创建HTTP响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            # 保存工作簿到响应
            wb.save(response)

            logger.info(f"Excel导出成功: {filename}, 记录数: {count}, 字段数: {len(valid_fields)}")
            return response

        except Exception as e:
            logger.error(f"Excel导出失败: {str(e)}")
            raise

    @staticmethod
    def _get_field_value(record, field):
        """根据字段名获取记录的值"""
        try:
            if field == 'match_status':
                return record.get_match_status_display()
            elif field == 'matched_at' and getattr(record, field):
                return getattr(record, field).strftime('%Y-%m-%d %H:%M:%S')
            elif field == 'created_at':
                return getattr(record, field).strftime('%Y-%m-%d %H:%M:%S')
            else:
                value = getattr(record, field)
                return value if value is not None else ''
        except Exception as e:
            logger.warning(f"获取字段 {field} 的值失败: {str(e)}")
            return ''

    @staticmethod
    def _auto_adjust_columns(worksheet):
        """自动调整列宽"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

    @staticmethod
    def export_with_filters(queryset, filters=None, export_all=False, fields=None):
        """
        根据过滤条件导出数据 - 支持字段选择

        Args:
            queryset: 基础查询集
            filters: 过滤条件字典
            export_all: 是否导出所有数据（忽略分页）
            fields: 要导出的字段列表

        Returns:
            HttpResponse: Excel文件响应
        """
        try:
            # 应用过滤条件
            filtered_queryset = queryset

            if filters:
                # 处理匹配状态过滤
                if 'match_status' in filters and filters['match_status']:
                    filtered_queryset = filtered_queryset.filter(match_status=filters['match_status'])

                # 处理车牌号码过滤
                if 'vehicle_number' in filters and filters['vehicle_number']:
                    filtered_queryset = filtered_queryset.filter(vehicle_number__icontains=filters['vehicle_number'])

                # 处理车主姓名过滤
                if 'owner_name' in filters and filters['owner_name']:
                    filtered_queryset = filtered_queryset.filter(owner_name__icontains=filters['owner_name'])

                # 处理VIN码过滤
                if 'vin' in filters and filters['vin']:
                    filtered_queryset = filtered_queryset.filter(vin__icontains=filters['vin'])

                # 处理品牌过滤
                if 'brand' in filters and filters['brand']:
                    filtered_queryset = filtered_queryset.filter(brand__icontains=filters['brand'])

                # 处理使用性质过滤
                if 'use_character' in filters and filters['use_character']:
                    filtered_queryset = filtered_queryset.filter(use_character__icontains=filters['use_character'])

                # 处理燃料种类过滤
                if 'energy_type' in filters and filters['energy_type']:
                    filtered_queryset = filtered_queryset.filter(energy_type__icontains=filters['energy_type'])

            # 如果选择导出所有数据，忽略分页
            if export_all:
                filtered_queryset = filtered_queryset.order_by('-created_at')
            else:
                # 默认只导出当前页的数据
                filtered_queryset = filtered_queryset.order_by('-created_at')

            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filter_info = ""

            if filters and any(filters.values()):
                filter_parts = []
                if filters.get('match_status'):
                    filter_parts.append(filters['match_status'])
                if filters.get('vehicle_number'):
                    filter_parts.append(f"车牌{filters['vehicle_number']}")
                if filters.get('owner_name'):
                    filter_parts.append(f"车主{filters['owner_name']}")

                if filter_parts:
                    filter_info = "_" + "_".join(filter_parts)

            filename = f"报废车信息{filter_info}_{timestamp}.xlsx"

            return ExcelExporter.export_scrap_car_info(filtered_queryset, filename, fields=fields)

        except Exception as e:
            logger.error(f"带过滤条件的Excel导出失败: {str(e)}")
            raise

    @staticmethod
    def get_available_fields():
        """获取所有可用的导出字段"""
        return ExcelExporter.FIELD_MAPPING

    @staticmethod
    def get_default_fields():
        """获取默认导出字段"""
        return ExcelExporter.DEFAULT_FIELDS
